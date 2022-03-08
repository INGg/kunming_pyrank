import os
import pprint

import openpyxl

# path = r"C:\Users\asuka\Desktop"
# os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('46届ICPC(昆明)中文问卷报名统计表.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook.active  # 获取活动表
print('当前活动表是：' + str(sheet))

# 按行获取值
print('按行获取值')

team = {
    0: [],
    1: [],
    2: [],
    3: []
}

for i in sheet.iter_rows(min_row=2, max_row=884, min_col=1, max_col=26):
    dayi = 0
    s = ""
    for j in i:
        s += " " + str(j.value)
        if str(j.value).strip() == '大一':
            dayi += 1
        # print(j.value, end=' ')
    # print()

    team[dayi].append(s)

print(len(team[0]))
pprint.pprint(team[3])

# # 按列获取值
# print('按列获取值')
# for i in sheet.iter_cols(min_row=2, max_row=5, min_col=1, max_col=2):
#     for j in i:
#         print(j.value)

# 结果：
# 当前活动表是：<Worksheet "Sheet1">
# 按行获取值
# 张三
# 74
# 李四
# 41
# 王五
# 56
# 赵六
# 12
# 按列获取值
# 张三
# 李四
# 王五
# 赵六
# 74
# 41
# 56
# 12
